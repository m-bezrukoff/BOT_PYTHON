from simulator_v2.playground_enter_points import *
from inc.inc_system import *
import openpyxl
from openpyxl.styles import Alignment, Font


ft9 = Font(size=9)
v_al = Alignment(horizontal="right", vertical="center")


def get_wb(file):
    return openpyxl.load_workbook(filename=file)


def get_ws(wb, page):
    return wb[page]


def update_sheets_data(wb, data):
    ws = wb['test']
    row = ws.max_row + 1

    allocation = {
        'total_percent': 1,
        'total_money': 2,
        'wins': 3,
        'losts': 4,
        'ratio': 5,
        'ema_f': 6,
        'ema_f_phase': 7,
        'ema_f_power': 8,
        'ema_m': 9,
        'ema_m_phase': 10,
        'ema_m_power': 11,
    }

    for key in data:
        c = ws.cell(row=row, column=allocation[key], value=data[key])
        c.font = ft9
        c.alignment = v_al


def xls_update(file, data):
    print('barbers in database:', len(data))
    wb = get_wb(file)
    update_sheets_data(wb, data)
    wb.save(file)


class Summary:
    def __init__(self, data):
        self.params = data.params
        self.result = data.result
        self.tradebook = data.data.tradebook
        self.timeline = [i['date'] for i in self.tradebook]
        self.consolidated = self.render_cosolidated_tradebook()
        self.points = [i for i in self.result if i['points_open'] or i['points_close']]
        self.summary = []
        self.positions = []
        self.xls_result = {}

        self.calculate_positions()
        self.calcutale_total()

    def render_cosolidated_tradebook(self):
        consolidated = {int(i['date']): {'trade': i} for i in self.tradebook}
        for i in self.result:
            if i['points_open']:
                date = i['points_open']['trade_timestamp']
                consolidated[date]['points_open'] = i['points_open']
            if i['points_close']:
                date = i['points_close']['trade_timestamp']
                consolidated[date]['points_close'] = i['points_close']
        return consolidated

    def calculate_positions(self):
        frame = self.params['frame']
        position = {}
        is_position = False
        for i in self.consolidated:
            if not is_position:
                if self.consolidated[i].get('points_open'):     # открываем позицию
                    is_position = True
                    position = dict()
                    position['timestamp_open'] = self.consolidated[i]['points_open']['trade_timestamp']
                    position['candle_timestamp'] = self.consolidated[i]['points_open']['candle_timestamp']
                    position['rate_open'] = self.consolidated[i]['points_open']['y']
                    position['type'] = self.consolidated[i]['points_open']['type']
                    position['stoploss'] = self.consolidated[i]['points_open']['stoploss']
                    position['utc_date_open'] = self.consolidated[i]['points_open']['x']
            else:
                if self.consolidated[i].get('points_close'):
                    is_position = False
                    position['timestamp_close'] = self.consolidated[i]['points_close']['trade_timestamp']
                    position['candle_timestamp'] = self.consolidated[i]['points_close']['candle_timestamp']
                    position['rate_close'] = self.consolidated[i]['points_close']['y']
                    position['utc_date_close'] = self.consolidated[i]['points_close']['x']
                    position['percent_profit'] = profit(position['rate_open'], position['rate_close'], 0.26, position['type'])
                    self.positions.append(position)
                    pass
                else:
                    # if self.consolidated[i]['trade']['rate'] > add_percent(position['rate_open'], 0.5):
                    #     is_position = False
                    #     position['timestamp_close'] = i
                    #     position['candle_timestamp'] = find_candle_start_time(i, frame)
                    #     position['rate_close'] = self.consolidated[i]['trade']['rate']
                    #     position['utc_date_close'] = utc_timestamp_to_date(position['candle_timestamp'])
                    #     position['percent_profit'] = profit(position['rate_open'], position['rate_close'], 0.26, position['type'])
                    #     self.positions.append(position)

                    if self.consolidated[i]['trade']['rate'] < position['stoploss']:
                        # print(self.consolidated[i]['trade'])
                        is_position = False
                        position['timestamp_close'] = i
                        position['candle_timestamp'] = find_candle_start_time(i, frame)
                        position['rate_close'] = self.consolidated[i]['trade']['rate']
                        position['utc_date_close'] = utc_timestamp_to_date(position['candle_timestamp'])
                        position['percent_profit'] = profit(position['rate_open'], position['rate_close'], 0.26, position['type'])
                        self.positions.append(position)

    def calcutale_total(self):
        print('-----------------------------------')
        percent_sum = 0
        total_positive = 0
        total_negative = 0
        initial_amount = 100
        final_amount = initial_amount

        for i in self.positions:
            percent_sum += i['percent_profit']
            final_amount = add_percent(final_amount, i['percent_profit'])
            if i['percent_profit'] > 0:
                total_positive += 1

            if i['percent_profit'] <= 0:
                total_negative += 1

            print('{} - {} {} {} {} ->: {}'.format(i['utc_date_open'], i['utc_date_close'], '\033[32m ' if i['percent_profit'] > 0 else '\033[31m', to2(i['percent_profit']), '%\033[0m', to2(final_amount)))
        print('-----------------------------------')
        amount_sum = percent(initial_amount, final_amount)
        amount_sum_text = '+' + to2(amount_sum) if amount_sum > 0 else to2(amount_sum)
        percent_sum_text = '+' + to2(percent_sum) if percent_sum > 0 else to2(percent_sum)

        print('{}total percent: {} % real money: {} %{}'.format('\033[32m' if percent_sum > 0 else '\033[31m', percent_sum_text, amount_sum_text, '\033[0m'))
        try:
            wr = (total_positive / (total_positive + total_negative)) * 100
        except ZeroDivisionError:
            wr = 0
        print('total wins: {}   losts: {}    win ratio: {} %'.format(total_positive, total_negative, to2(wr)))
        print('-----------------------------------')

        self.xls_result = {
            'total_percent': round(percent_sum, 2),
            'total_money': round(amount_sum, 2),
            'wins': int(total_positive),
            'losts': int(total_negative),
            'ratio': round(wr, 2),
            'ema_f': self.params['ema_f'],
            'ema_f_phase': self.params['ema_f_phase'],
            'ema_f_power': self.params['ema_f_power'],
            'ema_m': self.params['ema_m'],
            'ema_m_phase': self.params['ema_m_phase'],
            'ema_m_power': self.params['ema_m_power'],
        }

    def save_xls_results(self):
        if self.xls_result['ratio'] > 25 or self.xls_result['total_percent'] > 0:
            xls_file = 'reports/report.xlsx'
            # xls_new_from_template('inc/template.xlsx', xls_file)
            xls_update(xls_file, self.xls_result)


def profit(rate_in, rate_out, fee, typ):
    if typ == 'long':
        return percent(rate_in, rate_out) - fee
    elif typ == 'short':
        return -percent(rate_in, rate_out) - fee
    else:
        raise ValueError('type error in profit')
