import openpyxl
from openpyxl.styles import Alignment, Font


class Exporter:
    def __init__(self, file_path, page, save_interval):
        self.save_interval = save_interval
        self.path = file_path
        self.wb = openpyxl.load_workbook(filename=self.path)
        # xls_new_from_template('inc/template.xlsx', xls_file)
        self.ws = self.wb[page]
        self.row = self.ws.max_row + 1
        self.ft9 = Font(size=9)
        self.v_al = Alignment(horizontal="right", vertical="center")
        self.counter = 0

        self.allocation = {
            'total_percent': 1,
            'total_money': 2,
            'wins': 3,
            'losses': 4,
            'ratio': 5,
            'ema_f': 6,
            'ema_f_phase': 7,
            'ema_f_power': 8,
            'ema_m': 9,
            'ema_m_phase': 10,
            'ema_m_power': 11,
        }

    def update_xlx_results(self, result):
        for key in result:
            if key in self.allocation:
                c = self.ws.cell(row=self.row, column=self.allocation[key], value=result[key])
                c.font = self.ft9
                c.alignment = self.v_al
        self.row += 1
        self.counter += 1

        if self.counter % self.save_interval == 0:
            self.save_xls_file()

    def save_xls_file(self):
        self.wb.save(self.path)
